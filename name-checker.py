import re
import openpyxl

# Function to check if a string follows correct naming nomenclature
def check_naming_nomenclature(name):
    # Define your naming convention using regular expressions
    pattern = r'^[A-Za-z]+_[0-9]{4}$'
    return re.match(pattern, name)

# Load the Excel file with host names
input_file = 'names.xlsx'
wb = openpyxl.load_workbook(input_file)
sheet = wb.active

# Lists to store incorrect and correct names
incorrect_names = []
correct_names = []

# Iterate through the cells and check naming nomenclature
for row in sheet.iter_rows(min_row=2, values_only=True):
    name = row[0]
    if check_naming_nomenclature(name):
        correct_names.append(name)
    else:
        incorrect_names.append(name)

# Create a new Excel file to store incorrect names
output_file_incorrect = 'wrong_names.xlsx'
output_wb_incorrect = openpyxl.Workbook()
output_sheet_incorrect = output_wb_incorrect.active
output_sheet_incorrect.title = 'Incorrect Names'

# Write incorrect names to the new Excel file
for idx, name in enumerate(incorrect_names, start=1):
    output_sheet_incorrect.cell(row=idx, column=1, value=name)

# Save the new Excel file for incorrect names
output_wb_incorrect.save(output_file_incorrect)

# Write correct names to a text file
output_file_correct = 'correct_names.txt'
with open(output_file_correct, 'w') as f:
    for name in correct_names:
        f.write(name + '\n')

# Close the input and output workbooks
wb.close()
output_wb_incorrect.close()

print("Incorrect names extracted and saved to", output_file_incorrect)
print("Correct names extracted and saved to", output_file_correct)
